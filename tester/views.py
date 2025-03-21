import os
import time
from datetime import datetime

import pandas as pd
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse, HttpResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt

from .forms import *
from django.contrib.auth.models import Group
from django.contrib.auth.models import Permission
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.core.paginator import Paginator
from reversion.models import Version
from openpyxl import load_workbook
import telebot
from telebot import types
from . import cora_bot
from django.core.exceptions import PermissionDenied


bot = telebot.TeleBot('6924477556:AAH3pYP8AzQJXia27bgxAW1srTfAAaCaHC0')

# TOKEN = '6924477556:AAH3pYP8AzQJXia27bgxAW1srTfAAaCaHC0'
# tbot = telebot.TeleBot(TOKEN)
#
# def set_webhook(request):
#     bot_info = tbot.get_me()
#     webhook_url = f"http://127.0.0.1:8000/set_webhook/"  # замените на ваш URL-адрес
#     result = tbot.set_webhook(url=webhook_url)
#     if result:
#         return JsonResponse({'status': 'Webhook has been set'})
#     else:
#         return JsonResponse({'status': 'Failed to set webhook'})
#
# @csrf_exempt
# def bot(request):
#     if request.method == 'POST':
#         json_data = request.body.decode('utf-8')
#         update = telebot.types.Update.de_json(json_data)
#         tbot.process_new_updates([update])
#         return HttpResponse('')
#     else:
#         return HttpResponseNotAllowed(['POST'])
#
#
# @tbot.message_handler(content_types=["text"])
# def get_okn(message):
#     tbot.send_message(message.chat.id, "Hello, bot!")


def start_page(request):
    if request.user.is_authenticated:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            proc = Ticket.objects.get_queryset_none().filter(viewed=False).values('id', 'district', 'street', 'house',
                                                                                  'apartment',
                                                                                  'date', 'date_change', 'closed_date',
                                                                                  'completion_date', 'login',
                                                                                  'first_contact',
                                                                                  'second_contact', 'comment_master',
                                                                                  'comment_operator', 'operator__id',
                                                                                  'operator__username', 'master__id',
                                                                                  'master__username', 'type',
                                                                                  'priority', 'status',
                                                                                  'cause', 'user_change', 'viewed',
                                                                                  'deleted')
            user_per = list(request.user.get_group_permissions())

            data = {'tickets': list(proc), 'user_permissions': user_per}
            return JsonResponse(data, safe=False)

        if request.method == 'POST':
            form = NewsForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request,
                                 f'Новость добавлена')
                redirect('start_page')
            else:
                messages.error(request, f'Новость не может быть добавлена {Exception(request)}')

            return redirect('start_page')

        elif request.method == 'GET':
            username = request.user.get_username()
            ticket_time = timezone.now()
            is_super = bool(request.user.is_superuser)
            get_mater_ticket = TicketFilterForm(request.GET,queryset=Ticket.objects.get_master(id=request.user.id).order_by( "-date"), user=request.user)
            news = News.objects.all().order_by("-date")
            news_form = NewsForm()
            change_master = TicketForm()
            tickets = TicketFilterForm(request.GET, queryset=Ticket.objects.all())
            paginator = Paginator(tickets.qs, 25)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            context = {
                "tickets": tickets,
                "page_obj": page_obj,
                "username": username,
                "is_super": is_super,
                "master_ticket": get_mater_ticket,
                "change_master": change_master,
                "news": news,
                "n_form": news_form,
                "time": ticket_time,
                "pages": paginator.page_range,

            }

            return render(request, 'tester/start_page.html', context)

    else:
        return redirect("login")


def login_in_telegram(request):
    if request.method == 'POST':
        messages.success(request, f'Вы авторизовались в телерам как мастер {request.user.first_name}')
        return redirect('profile')

    return render(request, 'tester/login_in_telegram.html')


@permission_required('tester.master', login_url='error')
def add_com_master(request, pk):
    if request.method == 'POST':
        try:
            instance = Ticket.objects.get(id=pk)
            com_master_edit(request, instance)


        except:
            messages.error(request, f'Данные не могут быть изменены {Exception()}')

    username = request.user.get_username()
    year_now = timezone.now()
    on = Ticket.objects.get(id=pk)
    edit_from = AddComMaster(instance=on)
    context = {
        "username": username,
        "form": edit_from,
        "year_now": year_now
    }
    return render(request, 'tester/add_comment_master.html', context)


@permission_required('tester.operator', login_url='error')
def add_ticket(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            instance = None
            form = TicketForm(request.POST, instance=instance)
            if form.is_valid():
                ticket = form.save(commit=False)
                ticket.operator = request.user
                ticket.status = None
                ticket.comment_master = None
                ticket.user_change = request.user.first_name
                ticket.save()
                master_message(request, ticket)
                if ticket.id:
                    return redirect('edit_ticket', ticket.id)
            else:
                messages.error(request, "Ошибка при сохранении формы. Проверьте данные.")

        else:
            form = TicketForm()

    news = News.objects.all()
    ticket_time = timezone.now()
    username = request.user.get_username()
    year_now = timezone.now()
    proc = Ticket.objects.get_queryset_none()

    context = {
        "form": form,
        "username": username,
        "year_now": year_now,
        "news": news,
        "time": ticket_time,
        "processing": proc,
    }
    return render(request, 'tester/add_ticket.html', context)


@permission_required('tester.master', login_url='error')
def com_master_edit(request, instance=None):
    form = AddComMaster(request.POST, instance=instance)
    if form.is_valid():
        ticket = form.save(commit=False)
        ticket.user_change = request.user.get_username()
        ticket.viewed = False
        ticket.save()

        if ticket.cause == False:
            messages.success(request, f'Заявка {instance.street}  {instance.house} закрыта по пречине недозвон')
            ticket.closed_date = timezone.now()
            ticket.completion_date = None
            ticket.save()


        if ticket.cause == None:
            messages.success(request,f'Заявка {instance.street}  {instance.house} выполнена')
            ticket.completion_date = timezone.now()
            ticket.closed_date = None
            ticket.save()


        return redirect('master_comment', ticket.id)

    else:
        messages.error(request, f'Данные не могут быть изменены {Exception(request)}')


def login_cora_2(request):
    year_now = timezone.now()
    context = {
        "year_now": year_now
    }
    if request.method == 'POST':
        username = request.POST.get('login')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            messages.success(request, 'Вы вошли в систему')
            login(request, user)
            return redirect('start_page')

        else:
            messages.error(request, 'Логин или пароль введеный не верно')
            redirect('login')

    return render(request, 'tester/login_form.html', context=context)


def logout_cora_2(request):
    logout(request)
    return redirect('login')


def profile(request):
    username = request.user.get_username()
    year_now = timezone.now()
    user = User.objects.get(pk=request.user.id)
    password = request.POST.get('password')
    password_confirm = request.POST.get('password_confirm')
    first_name = request.POST.get('first_name')

    if first_name:
        user.first_name = first_name
        user.save()

    if password:
        if password == password_confirm:
            user.set_password(password)
            messages.success(request, 'Пароль обновлен')

        else:
            messages.error(request, 'Пароли не совпадают, текущий пароль не изменен')

    user.save()
    context = {
        "user": user,
        "username": username,
        "time": year_now
    }
    return render(request, 'tester/profile.html', context)


@permission_required('tester.dispatcher', login_url='error')
def processing(request):
    if request.user.is_authenticated:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            proc = Ticket.objects.get_queryset_none().filter(viewed=False).values('id', 'district', 'street', 'house',
                                                                                  'apartment',
                                                                                  'date', 'date_change', 'closed_date',
                                                                                  'completion_date', 'login',
                                                                                  'first_contact',
                                                                                  'second_contact', 'comment_master',
                                                                                  'comment_operator', 'operator__id',
                                                                                  'operator__username', 'master__id',
                                                                                  'master__username', 'type',
                                                                                  'priority', 'status',
                                                                                  'cause', 'user_change', 'viewed',
                                                                                  'deleted')
            data = list(proc)
            return JsonResponse(data, safe=False)

    news = News.objects.all()
    ticket_time = timezone.now()
    change_master = TicketForm()
    tickets = TicketFilterFormProcessing(request.GET, queryset=Ticket.objects.get_queryset_none())
    username = request.user.get_username()
    proc = Ticket.objects.filter(status=None)
    is_super = bool(request.user.is_superuser)

    paginator = Paginator(tickets.qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "tickets": tickets,
        "username": username,
        "is_super": is_super,
        "change_master": change_master,
        "news": news,
        "time": ticket_time,
        "processing": proc,
        "page_obj": page_obj,
        "pages": paginator.page_range,

    }
    return render(request, 'tester/processing.html', context)


def master_message(request, obj):
    bot = telebot.TeleBot('6924477556:AAH3pYP8AzQJXia27bgxAW1srTfAAaCaHC0')
    master_telegram_id = obj.master.telegram_id
    if master_telegram_id:
        if obj.apartment:
            bot.send_message(master_telegram_id, f"Открыта новая заявка {obj.street} {obj.house} кв {obj.apartment}")
        else:
            bot.send_message(master_telegram_id, f"Открыта новая заявка {obj.street} {obj.house}")
    else:
        messages.error(request, "Не удалось отправить уведомление: отсутствует telegram_id для мастера.")


@permission_required('tester.operator', login_url='error')
def edit_ticket(request, pk):
    if request.method == "GET":
        if request.user.has_perm("tester.dispatcher"):
            form = Ticket.objects.get(int(pk))
            form.viewed = True
            form.save()
    else:
        pass

    if request.method == 'POST':
        try:
            instance = Ticket.objects.get(id=pk)
            form = TicketForm(request.POST, instance=instance)
            if form.is_valid():
                ticket = form.save(commit=False)
                ticket.viewed = False
                ticket.user_change = request.user.first_name
                ticket.date_change = timezone.now()
                ticket.save()

                if ticket.cause == False:
                    messages.success(request, f'Заявка {instance.street}  {instance.house} закрыта по пречине недозвон')
                    ticket.closed_date = timezone.now()
                    ticket.completion_date = None
                    ticket.save()

                if ticket.cause is None:
                    messages.success(request, f'Заявка {instance.street}  {instance.house} выполнена')
                    ticket.completion_date = timezone.now()
                    ticket.closed_date = None
                    ticket.save()

                if ticket.status:
                    ticket.deleted = False
                    master_message(request, ticket)
                    ticket.save()

                elif not ticket.status:
                    ticket.deleted = True
                    ticket.save()

                elif ticket.status is None:
                    ticket.viewed = False
                    ticket.save()

                if instance.apartment is None:
                    messages.success(request,f'Данные успешно изменены {instance.street}  {instance.house}')

                else:
                    messages.success(request, f'Данные успешно изменены {instance.street}  {instance.house} кв {instance.apartment}')

                return redirect('edit_ticket', pk)

        except:
            messages.error(request, f'Данные не могут быть изменены {Exception(request)}')

    versions = Version.objects.get_for_object(Ticket.objects.get(id=pk))

    changes_list = []

    for version in versions:
        changes = version.field_dict.items()

        changed_fields = []
        for field, value in changes:
            changed_fields.append({
                'field': field,
                'old_value': None,
                'new_value': value,
            })

        changes_list.append({
            'version_id': version.id,
            'changed_fields': changed_fields,
        })

    get_odj = Ticket.objects.get(id=pk)
    username = request.user.get_username()
    year_now = timezone.now()
    on = Ticket.objects.get(id=pk)
    edit_from = TicketForm(instance=on)
    proc = Ticket.objects.get_queryset_none()

    context = {
        "e_form": edit_from,
        "username": username,
        "time": year_now,
        "obj": get_odj,
        "processing": proc,
        'changes_list': changes_list,
    }

    return render(request, 'tester/edit_ticketfrom.html', context)


def error(request):
    username = request.user.get_username()
    year_now = timezone.now()
    messages.error(request, 'Недостаточно прав')
    context = {
        "username": username,
        "time": year_now,
    }
    return render(request, 'tester/error.html', context)


def log(request):
    username = request.user.get_username()

    context = {
        "username": username,
    }
    return render(request, 'tester/log.html', context)


def handle_uploaded_file(file):
    # Создайте отдельный путь для загруженных файлов
    upload_path = os.path.join(settings.MEDIA_ROOT, 'uploads')
    os.makedirs(upload_path, exist_ok=True)

    # Сохраните файл в новый путь
    fs = FileSystemStorage(location=upload_path)
    filename = fs.save(file.name, file)

    # Загрузите данные Excel с использованием pandas
    excel_path = os.path.join(upload_path, filename)
    df = pd.read_excel(excel_path)

    # Добавьте вашу логику обработки данных здесь

    # Верните данные в представление (здесь просто для примера)
    return render(None, 'tester/shutdown.html', {'data': df.to_html()})


@permission_required('tester.dispatcher', login_url='error')
def shutdown(request):
    if request.method == 'POST':
        wdw = ShutdownForm(request.POST, request.FILES)
        if wdw.is_valid():
            excel_file = request.FILES['file']
            try:
                wb = load_workbook(excel_file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    street, house, apartment = row
                    master_id = request.POST.get('master')
                    obj = Shutdown(street=street, house=house, apartment=apartment, master_id=master_id)
                    obj.save()

                messages.success(request, 'Отключения успешно добавлены.')
            except Exception as e:
                messages.error(request,f'Ошибка при обработке файла: {e}')  # Вывод ошибки для проверки

        else:
            messages.error(request, 'Форма недействительна. Пожалуйста, проверьте введенные данные.')
            messages.error(request, wdw)

    username = request.user.get_username()
    form = ShutdownForm()
    proc = Ticket.objects.get_queryset_none()
    year_now = timezone.now()
    master_shutdown = ShutdownFilter(request.GET, queryset=Shutdown.objects.all())

    context = {
        "master": master_shutdown,
        "username": username,
        "form": form,
        "processing": proc,
        "time": year_now,
    }
    return render(request, 'tester/shutdown.html', context)


@permission_required('tester.dispatcher', login_url='error')
def del_shutdown(request, pk):
    get_obj = Shutdown.objects.get(id=pk)
    get_obj.delete()

    if get_obj.apartment:
        messages.success(request, f"Отключка {get_obj.street} {get_obj.house} кв {get_obj.apartment} удалена")
    else:
        messages.success(request, f"Отключка {get_obj.street} {get_obj.house} удалена")

    return redirect('shutdown')


@permission_required('tester.master', login_url='error')
def shutdown_master(request):
    if request.method == 'POST':
        completed_ids = request.POST.getlist('completion_checkbox')
        if completed_ids:
            Shutdown.objects.filter(id__in=completed_ids).update(completion=True)
            Shutdown.objects.filter(id__in=completed_ids).update(closed_date=timezone.now())
            Shutdown.objects.filter(id__in=completed_ids).update(deleted=True)
            messages.success(request, 'Оключка(и) выполнена(ы)')

    username = request.user.get_username()
    master_shutdown = ShutdownFilterMaster(request.GET, queryset=Shutdown.objects.filter(master_id=request.user.id))
    form = Shutdownlist()
    is_super = bool(request.user.is_superuser)
    for_super = ShutdownFilterMaster(request.GET, queryset=Shutdown.objects.all())
    year_now = timezone.now()
    button = Shutdown.objects.all().filter(master_id=request.user.id)
    context = {
        "for_super": for_super,
        "is_super": is_super,
        "form": form,
        "username": username,
        "master": master_shutdown,
        "time": year_now,
        "button": button,
    }
    return render(request, 'tester/shutdown_master.html', context)


def include_master(request):
    if request.method == 'POST':
        completed_ids = request.POST.getlist('completion_checkbox')
        if completed_ids:
            Shutdown.objects.filter(id__in=completed_ids).update(completion=False)
            Shutdown.objects.filter(id__in=completed_ids).update(closed_date=None)
            Shutdown.objects.filter(id__in=completed_ids).update(deleted=False)
            messages.success(request, 'Оключка(и) снова активна(ы)')

    username = request.user
    master_shutdown = ShutdownFilterMaster(request.GET, queryset=Shutdown.objects.filter(master_id=request.user.id).filter(completion=True))
    form = Shutdownlist()
    is_super = bool(request.user.is_superuser)
    for_super = ShutdownFilterMaster(request.GET, queryset=Shutdown.objects.all().filter(completion=True))
    year_now = timezone.now()
    context = {
        "for_super": for_super,
        "is_super": is_super,
        "form": form,
        "username": username,
        "master": master_shutdown,
        "time": year_now,
    }
    return render(request, 'tester/include_master.html', context)


@permission_required('tester.dispatcher', login_url='error')
def add_address(request):
    username = request.user.get_username()
    year_now = timezone.now()
    f = AddStreet()
    d = AddDistrict()
    street = Street.objects.all()
    district = District.objects.all()
    proc = Ticket.objects.get_queryset_none()

    if request.method == "POST":
        street_form = AddStreet(request.POST)
        district_form = AddDistrict(request.POST)
        district_edit = request.POST.get('district_edit')
        street_edit = request.POST.get('street_edit')
        district_id = request.POST.get('district_id')
        street_id = request.POST.get('street_id')

        obj_district = District.objects.filter(id=district_id)
        obj_street = Street.objects.filter(id=street_id)

        for it in obj_district:
            if district_edit == it.name:
                District.objects.filter(id=district_id).delete()
                messages.success(request, f'{district_edit} район был изменён')

            else:
                District.objects.filter(id=district_id).update(name=district_edit)
                messages.success(request, f'Название района было изменено с {it.name} на {district_edit}')

        for it in obj_street:
            if street_edit == it.name:
                Street.objects.filter(id=street_id).delete()
                messages.success(request, f'Улица {street_edit} была удалена ')

            else:
                Street.objects.filter(id=street_id).update(name=street_edit)
                messages.success(request, f'Название улицы было изменено с {it.name} на {street_edit}')

        if district_form.is_valid():
            district_form.save()
            messages.success(request, 'Район добавлен')

        elif street_form.is_valid():
            street_form.save()
            messages.success(request, 'Улица добавлена')

    context = {
        "username": username,
        "time": year_now,
        "form_street": f,
        "form_district": d,
        "streets": street,
        "districts": district,
        "processing": proc,
    }
    return render(request, 'tester/address.html', context)


@permission_required('tester.dispatcher', login_url='error')
def territory(request):
    username = request.user.get_username()
    year_now = timezone.now()
    proc = Ticket.objects.get_queryset_none()
    context = {
        "username": username,
        "time": year_now,
        "processing": proc,
    }
    return render(request, 'tester/territory.html', context)


@permission_required('tester.dispatcher', login_url='error')
def edit_news(request, pk):
    if request.method == 'POST':
        instance = News.objects.get(id=pk)
        form = NewsForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Новость успешно изменена ')

            return redirect('start_page')

        else:
            messages.success(request,'Новость не может быть изменена')

            return redirect('start_page')

    news = News.objects.all().order_by("-date")
    news_pk = News.objects.get(id=pk)
    news_form = NewsForm(instance=news_pk)
    year_now = timezone.now()

    context = {
        "time": year_now,
        'news': news,
        'news_pk': news_pk,
        'news_form': news_form,
    }

    return render(request, 'tester/edit_news.html', context)


@permission_required('tester.dispatcher', login_url='error')
def info(request):
    status = [None, True, False]
    types = ['Ремонт','Настройка','Перенос','Включение','Отключение','Установка','Тюнер']
    priority = ['Обычный','Срочный','Корпоративный']
    data = {}
    data2 = {}
    data3 = {}
    data4 = {}
    data5 = {}
    for const in status:
        obj = Ticket.objects.filter(status=const)
        data[f'{const}'] = len(obj)

    for it in types:
        obj2 = Ticket.objects.filter(type=it)
        data2[f'{it}'] = len(obj2)

    for caus in status:
        obj = Ticket.objects.filter(cause=caus)
        data3[f'{caus}'] = len(obj)

    for pri in priority:
        obj4 = Ticket.objects.filter(priority=pri)
        data4[f'{pri}'] = len(obj4)

    obj5 = Shutdown.objects.all()
    data5[f'Отключки'] = len(obj5)

    year_now = timezone.now()
    tickets = TicketFilterForm(request.GET, queryset=Ticket.objects.all())

    context = {
        "tickets": tickets,
        "data": data,
        "data2": data2,
        "data3": data3,
        "data4": data4,
        "data5": data5,
        "time": year_now,
    }
    return render(request, 'tester/info.html', context)


@permission_required('tester.dispatcher', login_url='error')
def delete_obj(request, pk):
    post_to_delete = News.objects.get(id=pk)
    messages.success(request, f"Новость {post_to_delete.title} удалена")
    post_to_delete.delete()
    return redirect('start_page')


def test(request):
    return render(request, 'tester/test.html')





